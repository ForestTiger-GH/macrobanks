# -*- coding: utf-8 -*-
"""
Единый конвейер для данных ЦБ РФ по финансированию долевого строительства (счета эскроу).

Шаги:
1) fetch_excel_links -> ссылки на ежемесячные .xlsx
2) download_and_parse_excels -> объединённый "длинный" поток данных (Регион, Показатель, Значение, Дата)
3) build_pivots -> набор сводных таблиц по каждому показателю (регионы × даты)
4) build_workbook -> оформление Excel (каждый показатель на отдельном листе)
5) save_workbook / download_file_if_colab -> сохранить файл / скачать в Colab

Поддержан макроаргумент INBANK:
- INBANK=False  -> прямые внешние URL cbr.ru
- INBANK=True   -> все URL прокидываются через apply_inbank_prefix(INBANK_API_PREFIX, url)

Зависимости (pip):
    numpy, pandas, requests, beautifulsoup4, tqdm, openpyxl
"""

from __future__ import annotations

import os
import re
import io
import shutil
import warnings
from tempfile import NamedTemporaryFile
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Импорт универсального префикса для внутренних шлюзов
# (Этот модуль должен быть в твоём репозитории: macrobanks/codes.py)
from macrobanks.routines import apply_inbank_prefix


# ---------------------------
# Настройки и константы
# ---------------------------

CBR_INDEX_URL_DEFAULT = "https://www.cbr.ru/statistics/bank_sector/equity_const_financing/"
CBR_BASE = "https://www.cbr.ru"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}

# Оформление Excel
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="000000")
DATA_FONT = Font(name="Arial", size=10)
FIRST_COL_FONT = Font(name="Arial", size=10, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
PALE_GREEN_FILL = PatternFill("solid", fgColor="D8F5DC")
BORDER_STYLE = Side(border_style="thin", color="D3D3D3")
CELL_BORDER = Border(top=BORDER_STYLE, bottom=BORDER_STYLE, left=BORDER_STYLE, right=BORDER_STYLE)
WIDTH_FIRST_COL = 25
WIDTH_OTHER_COLS = 11


# ---------------------------
# Вспомогательные функции URL/HTTP
# ---------------------------

def _effective_url(url: str, *, INBANK: bool, INBANK_API_PREFIX: str) -> str:
    """Возвращает URL с учётом INBANK и префикса."""
    if not INBANK:
        return url
    return apply_inbank_prefix(INBANK_API_PREFIX=INBANK_API_PREFIX, original_url=url)


def _http_get(url: str, *, session: Optional[requests.Session] = None, **kwargs) -> requests.Response:
    """GET с дефолтными заголовками и проверкой статуса."""
    sess = session or requests.Session()
    resp = sess.get(url, headers=HEADERS, timeout=60, **kwargs)
    resp.raise_for_status()
    return resp


# ---------------------------
# 1) Получение списка ссылок на Excel
# ---------------------------

def fetch_excel_links(
    *,
    INBANK: bool = True,
    INBANK_API_PREFIX: str = "https://gate.rshb.ru/ex/cbr",
    index_url: str = CBR_INDEX_URL_DEFAULT,
    session: Optional[requests.Session] = None,
) -> List[str]:
    """Возвращает список абсолютных ссылок на ежемесячные .xlsx."""
    url = _effective_url(index_url, INBANK=INBANK, INBANK_API_PREFIX=INBANK_API_PREFIX)
    html = _http_get(url, session=session).text
    soup = BeautifulSoup(html, "html.parser")

    links: List[str] = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.lower().endswith(".xlsx"):
            if not href.startswith("http"):
                href = CBR_BASE + href
            links.append(href)

    # Убираем дубликаты и прокидываем префикс на каждый файл (если INBANK)
    links = sorted(set(_effective_url(h, INBANK=INBANK, INBANK_API_PREFIX=INBANK_API_PREFIX) for h in links))
    return links


# ---------------------------
# 2) Парсинг одного файла и объединение "потока"
# ---------------------------

_DATE_RE = re.compile(r"(\d{2})(\d{2})(\d{4})")  # DDMMYYYY

def _extract_date_from_filename(name: str) -> Optional[str]:
    """Извлекает дату вида YYYY-MM-DD из имени файла (шаблон DDMMYYYY)."""
    m = _DATE_RE.search(name)
    if not m:
        return None
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"


def clean_indicator_name(s: str) -> str:
    """Удаляет хвостовые сноски/цифры/звёздочки."""
    s = str(s)
    s = re.sub(r"[\s\*\d]+$", "", s).rstrip()
    return s


def parse_cbr_equity_file(file_path: str) -> pd.DataFrame:
    """
    Читает один .xlsx и возвращает "длинные" данные:
    колонки: Регион, Показатель, Значение, Дата
    """
    date_str = _extract_date_from_filename(os.path.basename(file_path))
    df = pd.read_excel(file_path, header=3)

    # 1-й столбец — номера/округа; 2-й — регионы (иногда пустые для ФО)
    region_col = df.columns[1]
    df[region_col] = df[region_col].fillna(df.iloc[:, 0])
    df = df.iloc[:, 1:]  # удаляем самый левый столбец

    long_df = df.melt(id_vars=[region_col], var_name="Показатель", value_name="Значение")
    long_df = long_df.rename(columns={region_col: "Регион"})
    long_df["Дата"] = date_str
    long_df = long_df.dropna(subset=["Значение"])

    # Очистки как в исходном коде
    long_df["Показатель"] = long_df["Показатель"].apply(clean_indicator_name)
    long_df["Регион"] = (
        long_df["Регион"]
        .astype(str)
        .apply(lambda x: re.sub(r"\d+$", "", x).strip())
        .replace("Итого", "Итого по РФ")
    )
    return long_df


def download_and_parse_excels(
    links: List[str],
    *,
    workdir: str = "temp_equity_xlsx",
    session: Optional[requests.Session] = None,
    verbose: bool = True,
) -> pd.DataFrame:
    """Скачивает все ссылки (если нужно) и объединяет в один поток данных."""
    os.makedirs(workdir, exist_ok=True)
    frames: List[pd.DataFrame] = []

    iterator = tqdm(links, desc="Processing files") if verbose else links
    sess = session or requests.Session()

    for url in iterator:
        filename = url.split("/")[-1]
        local_path = os.path.join(workdir, filename)
        if not os.path.exists(local_path):
            with _http_get(url, session=sess, stream=True) as resp:
                with open(local_path, "wb") as f:
                    shutil.copyfileobj(resp.raw, f)

        try:
            frames.append(parse_cbr_equity_file(local_path))
        except Exception as e:
            warnings.warn(f"Ошибка обработки {filename}: {e}")

    if not frames:
        return pd.DataFrame(columns=["Регион", "Показатель", "Значение", "Дата"])

    result = pd.concat(frames, ignore_index=True)
    return result


# ---------------------------
# 3) Построение сводных таблиц (pivot)
# ---------------------------

def build_pivot(
    result: pd.DataFrame,
    indicator: str,
    *,
    full_regions_list: Optional[List[str]] = None,
    full_dates_list: Optional[List[str]] = None,
) -> pd.DataFrame:
    """Сводная матрица по одному показателю: регионы × даты."""
    last_date = result["Дата"].max()
    if full_regions_list is None:
        full_regions_list = result[result["Дата"] == last_date]["Регион"].unique().tolist()
    if full_dates_list is None:
        full_dates_list = sorted(result["Дата"].unique())

    df_ind = result[result["Показатель"] == indicator].copy()
    pivot = df_ind.pivot_table(index="Регион", columns="Дата", values="Значение")
    pivot = pivot.reindex(full_regions_list)
    pivot = pivot.reindex(columns=full_dates_list)
    return pivot


def _detect_indicators_order_from_latest_file(
    links: List[str],
    workdir: str,
) -> List[str]:
    """Определяем порядок столбцов показателей по последнему файлу (как в исходнике)."""
    # Берём ссылку с максимальной датой в имени
    def _key(u: str) -> str:
        m = re.findall(r"(\d{8})", u)
        return m[0] if m else ""

    if not links:
        return []

    last_link = sorted(links, key=_key, reverse=True)[0]
    last_local = os.path.join(workdir, last_link.split("/")[-1])
    if not os.path.exists(last_local):
        raise FileNotFoundError(
            f"Не найден последний файл {last_local}. Сначала вызови download_and_parse_excels()."
        )

    df_last = pd.read_excel(last_local, header=3)
    indicators = df_last.columns[2:].tolist()
    indicators = [clean_indicator_name(ind) for ind in indicators]
    return indicators


def build_pivots(
    result: pd.DataFrame,
    *,
    links: Optional[List[str]] = None,
    workdir: str = "temp_equity_xlsx",
) -> Tuple[Dict[str, pd.DataFrame], List[str], List[str], List[str]]:
    """
    Возвращает:
      pivots: {показатель -> DataFrame (регионы × даты)}
      indicators_order: порядок листов по последнему файлу
      full_regions_list, full_dates_list
    """
    full_dates_list = sorted(result["Дата"].unique())
    full_regions_list = result[result["Дата"] == result["Дата"].max()]["Регион"].unique().tolist()

    if links:
        indicators_order = _detect_indicators_order_from_latest_file(links, workdir=workdir)
    else:
        # Если ссылок нет, упорядочим по алфавиту
        indicators_order = sorted(result["Показатель"].unique())

    pivots: Dict[str, pd.DataFrame] = {}
    for ind in indicators_order:
        pivots[ind] = build_pivot(result, ind, full_regions_list=full_regions_list, full_dates_list=full_dates_list)

    return pivots, indicators_order, full_regions_list, full_dates_list


# ---------------------------
# 4) Построение и сохранение Excel
# ---------------------------

def _abbreviate_sheet_name(text: str) -> str:
    """Аббревиатура по первым буквам слов (только буквы и пробелы)."""
    text_clean = re.sub(r"[^а-яА-Яa-zA-Z ]", "", text)
    abbr = "".join(w[0].upper() for w in text_clean.split() if w)
    return abbr[:31] or "SHEET"  # ограничение Excel на 31 символ


def build_workbook(
    pivots: Dict[str, pd.DataFrame],
    indicators_order: List[str],
    *,
    title_row_font: Font = Font(name="Times New Roman", size=10, bold=True),
) -> Workbook:
    """Оформляет книгу Excel по заданным сводным таблицам."""
    wb = Workbook()
    wb.remove(wb.active)

    for indicator in tqdm(indicators_order, desc="Формирование листов Excel"):
        pivot_df = pivots[indicator].reset_index()
        ws = wb.create_sheet(title=_abbreviate_sheet_name(indicator))

        # Титул в A2
        ws["A2"] = indicator.upper()
        ws["A2"].font = title_row_font

        rows = list(dataframe_to_rows(pivot_df, index=False, header=True))
        start_row = 4
        for r_idx, row in enumerate(rows, start_row):
            highlight_row = False
            if r_idx > start_row:
                region_val = str(row[0])
                if " ФО" in region_val or "Итого по РФ" in region_val:
                    highlight_row = True

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == start_row:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(
                        horizontal="right" if c_idx > 1 else "left", vertical="center"
                    )
                    if c_idx == 1:
                        cell.font = FIRST_COL_FONT
                    if c_idx > 1 and isinstance(value, (int, float)) and pd.notnull(value):
                        cell.number_format = "#,##0"
                    if highlight_row:
                        cell.fill = PALE_GREEN_FILL

                cell.border = CELL_BORDER

        # Ширины, закрепление, сетка
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = WIDTH_FIRST_COL if col_idx == 1 else WIDTH_OTHER_COLS
        ws.freeze_panes = "B5"
        ws.sheet_view.showGridLines = False

    return wb


def save_workbook(wb: Workbook, out_path: str) -> str:
    """Сохраняет книгу на диск и возвращает путь."""
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def download_file_if_colab(out_path: str) -> None:
    """Если код запущен в Google Colab — инициируем загрузку файла в браузер."""
    try:
        from google.colab import files  # type: ignore
    except Exception:
        return
    try:
        files.download(out_path)
    except Exception:
        pass


# ---------------------------
# 5) Единая оболочка
# ---------------------------

def process_escrow_files(
    *,
    INBANK: bool = True,
    INBANK_API_PREFIX: str = "https://gate.rshb.ru/ex/cbr",
    index_url: str = CBR_INDEX_URL_DEFAULT,
    workdir: str = "temp_equity_xlsx",
    out_xlsx: str = "Escrow Accounts.xlsx",
    # auto_download_if_colab: bool = True,
    session: Optional[requests.Session] = None,
) -> str:
    """
    Полный цикл: ссылки -> поток -> pivots -> Excel -> (скачать в Colab).
    Возвращает путь к сохранённому .xlsx.
    """
    links = fetch_excel_links(
        INBANK=INBANK,
        INBANK_API_PREFIX=INBANK_API_PREFIX,
        index_url=index_url,
        session=session,
    )
    result = download_and_parse_excels(links, workdir=workdir, session=session)
    pivots, indicators_order, _, _ = build_pivots(result, links=links, workdir=workdir)
    wb = build_workbook(pivots, indicators_order)
    saved = save_workbook(wb, out_xlsx)
    # if auto_download_if_colab:
    #     download_file_if_colab(saved)
    return saved


# ---------------------------
# Экспорт API модуля
# ---------------------------

__all__ = ["process_escrow_files"]

# Вызов снаружи
# from cbr.escrow import process_escrow_files
# process_escrow_files(INBANK=True, INBANK_API_PREFIX="https://gate.ru/ex/cbr")
